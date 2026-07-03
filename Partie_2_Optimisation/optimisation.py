#un super code arrive soon



import pandas as pd
import re
import numpy as np
import datetime


STATUTS_DISPONIBLE = {"ASSU", "CP_REPORT", "DDD","DISPO","DISPO AM", "DISPO AMPL", "DISPO M", "DISPO MX","DISPO N"}
'''
def to_time(x):
    if isinstance(x, datetime.datetime):
        return x.time()
    elif isinstance(x, datetime.time):
        return x
    else:
        return datetime.datetime.strptime(str(x), "%H:%M").time()'''
def to_time(x):
    # Gestion des valeurs nulles / NaN / vides
    if pd.isna(x) or x is None or str(x).strip().lower() == 'nan':
        return datetime.time(0, 0) # Renvoie minuit par défaut pour éviter le plantage
        
    if isinstance(x, datetime.datetime):
        return x.time()
    elif isinstance(x, datetime.time):
        return x
    else:
        try:
            # Enlève les espaces inutiles autour du texte (ex: " 14:00 ")
            clean_str = str(x).strip()
            # Si le format Excel inclut les secondes (ex: "14:00:00")
            if len(clean_str.split(':')) == 3:
                return datetime.datetime.strptime(clean_str, "%H:%M:%S").time()
            return datetime.datetime.strptime(clean_str, "%H:%M").time()
        except ValueError:
            # En cas d'autre format texte imprévu, évite le crash global
            return datetime.time(0, 0)

def initialize_data(chemin_fichier_mach:str, chemin_fichier_serv, day: str):
    """
    Lit deux fichiers Excel :
    - un fichier contenant les machinistes ;
    - un fichier contenant les services.

    Le fichier des machinistes doit contenir au minimum :
    - une colonne "Identifiant" ;
    - une colonne "Qualification : Connaissance de ligne" ;
    - une colonne correspondant au jour donné dans l'argument day.

    Le fichier des services doit contenir au minimum :
    - une colonne "Service" ;
    - une colonne "Début" ;
    - une colonne "Fin".

    La fonction construit une matrice D de dimensions :
    nombre de machinistes × nombre de services.

    D[i, j] vaut 1 si le machiniste de la ligne i peut réaliser
    le service de la ligne j, et 0 sinon.

    Un machiniste peut réaliser un service si :
    - il connaît la ligne du service ;
    - et son statut du jour est compatible avec le service.

    Les statuts considérés comme disponibles sont :
    - une cellule vide / NaN ;
    - un statut appartenant à STATUTS_DISPONIBLE ;
    - "DISPO AM" si le service commence à partir de 14h00 ;
    - "DISPO M" si le service finit avant ou à 14h00 ;
    - "DISPO N" si le service finit après 22h ;
    - "DISPO" sans précision horaire.

    La fonction renvoie uniquement la matrice D.
    Elle ne renvoie pas les listes d'association entre identifiants réels
    et indices i ou j.
    """
    df_mach = (pd.read_excel(chemin_fichier_mach))[['Identifiant', 'Qualification : Connaissance de ligne', day]]
    df_serv = (pd.read_excel(chemin_fichier_serv)) [['Service', 'Début', 'Fin']]
    df_mach['qualification'] = df_mach['Qualification : Connaissance de ligne'].apply(
    lambda y: [int(x) for x in re.findall(r"\d+", str(y))] if pd.notna(y) else []
)
    N = len(df_mach['Identifiant'])
    P = len(df_serv['Service'])
    D = np.zeros((N,P), dtype=np.int8)
    for i in range(N):
        for j in range(P):
            service = df_serv['Service'].iloc[j]

            if pd.isna(service):
                continue

            ligne_service = int(str(service)[1:4])
            if ligne_service not in df_mach['qualification'].iloc[i]:
                continue
            if pd.isna(df_mach[day].iloc[i]) or (df_mach[day].iloc[i] in STATUTS_DISPONIBLE):
                D[i, j] = 1
            elif re.search("DISPO AM",df_mach[day].iloc[i]) and to_time(df_serv['Début'].iloc[j]) >= to_time(datetime.datetime.strptime("12:00", format="%H:%M")):
                D[i,j] = 1
            elif re.search("DISPO M",df_mach[day].iloc[i]) and to_time(df_serv['Fin'].iloc[j]) <= to_time(datetime.datetime.strptime("14:00", format="%H:%M")):
                D[i,j] = 1
            elif re.search("DISPO N",df_mach[day].iloc[i]) and (to_time(df_serv['Fin'].iloc[j]) >= to_time(datetime.datetime.strptime("22:00", format="%H:%M")) or to_time(df_serv['Fin'].iloc[j]) <= to_time(datetime.datetime.strptime("05:00", format="%H:%M"))):
                D[i,j] = 1
            elif re.search(r"\bDISPO\b(?!\s*(AM|M|N)\b)", str(df_mach[day].iloc[i])):
                D[i,j] = 1
    return D
            

D = initialize_data("Partie_1_LLM/data/Export_Planning_du_12_01_2026_au_16_01_2026.xlsx", 'Partie_1_LLM/data/Services_Agents_non_affectés_le_12_01_2026.xlsx', '12/01/2026')

def type_service(df, j):
    if (df['Type'].iloc[j]) != None and isinstance(df['Type'].iloc[j],str) :
        return df['Type'].iloc[j]
    elif to_time(df['Fin'].iloc[j]) <= to_time(datetime.datetime.strptime("14:00","%H:%M")):
        return 'MAT'
    elif to_time(df['Fin'].iloc[j]) >= to_time(datetime.datetime.strptime("22:00","%H:%M")) or to_time(df['Fin'].iloc[j]) <= to_time(datetime.datetime.strptime("05:00","%H:%M")):
        return 'NUIT'
    elif to_time(df['Début'].iloc[j]) >= to_time(datetime.datetime.strptime("14:00","%H:%M")):
        return 'AM'
    else:
        return 'JOUR'

def W_initialize(chemin_fichier_pref: str,chemin_fichier_serv: str, dim, D):
    """
    Initialise la matrice de pondération W associée aux affectations machinistes-services.

    Paramètres
    ----------
    chemin_fichier_pref : str
        Chemin vers le fichier Excel contenant les préférences des machinistes.
        Le fichier doit contenir au minimum :
        - une colonne "préférence_ligne" ;
        - une colonne "préférence_horaire".

        La colonne "préférence_ligne" contient les numéros de lignes préférées,
        par exemple "123, 456, 789".
        L'ordre des lignes est interprété comme un ordre de préférence :
        les premières lignes sont les plus préférées.

        La colonne "préférence_horaire" contient les types d'horaires préférés,
        par exemple "MAT, AM, JOUR, COUP, NUIT".
        L'ordre est également interprété comme un ordre de préférence.

    chemin_fichier_serv : str
        Chemin vers le fichier Excel contenant les services.
        Le fichier doit contenir au minimum :
        - une colonne "Service" ;
        - une colonne "Début" ;
        - une colonne "Fin" ;
        - une colonne "Type", ou les informations nécessaires à la fonction type_service.

    dim : tuple
        Dimensions de la matrice W sous la forme (N, P), où :
        - N est le nombre de machinistes ;
        - P est le nombre de services.

    D : np.ndarray
        Matrice de faisabilité de dimension (N, P).
        D[i, j] vaut 1 si le machiniste i peut faire le service j,
        et 0 sinon.

    Retour
    ------
    W : np.ndarray
        Matrice de pondération de dimension (N, P).
        Pour les affectations faisables, W[i, j] est ajustée selon :
        - les préférences de ligne du machiniste ;
        - les préférences horaires du machiniste.

        Les affectations non faisables ne sont pas modifiées dans W.
        Elles doivent donc être interdites séparément par la matrice D
        dans le modèle d'optimisation.
    """

    coef_ligne = 1
    coef_horaire = 3
    coef_existence = 12
    W = coef_existence * np.ones(dim)
    df_pref = pd.read_excel(chemin_fichier_pref)
    df_serv = pd.read_excel(chemin_fichier_serv)
    N, P = dim
    for i in range(N):
        liste_ligne = [int(x) for x in re.findall(r"\d+", str(df_pref['préférence_ligne'].iloc[i]))] if pd.notna(df_pref['préférence_ligne'].iloc[i]) else []
        liste_horaire = [x for x in re.findall(r"JOUR|AM|MAT|COUP|NUIT", str(df_pref['préférence_horaire'].iloc[i]))] if pd.notna(df_pref['préférence_horaire'].iloc[i]) else []
        increment_ligne = 1 / (len(liste_ligne) if len(liste_ligne) != 0 else 1)
        increment_horaire = 1 / (len(liste_horaire) if len(liste_horaire) != 0 else 1)
        for j in range(P):
            if D[i,j] == 0:
                continue
            index_ligne = liste_ligne.index(int(df_serv['Service'].iloc[j][1:4]))
            index_horaire = liste_horaire.index(type_service(df_serv, j))
            W[i,j] += coef_ligne * (0.5 + ((len(liste_ligne) if len(liste_ligne) != 0 else 1) - index_ligne) * increment_ligne) + coef_horaire * coef_horaire * (0.5 + (len(liste_horaire) - index_horaire) * increment_horaire)
    return W

            


W = W_initialize("Partie_2_Optimisation/preferences_agents.xlsx", 'Partie_1_LLM/data/Services_Agents_non_affectés_le_12_01_2026.xlsx', (len(D), len(D[0])), D)

#W=np.ones((len(D),len(D[0])))

#on considère qu'on a D et W la 
from ortools.linear_solver import pywraplp
def opti(W, D):
    
    costs=W
    num_workers=len(D)
    num_tasks=len(D[0])
    y=np.zeros((num_workers,num_tasks))
    solver= pywraplp.Solver.CreateSolver("SCIP")
    if not solver:
        return
    x={}
    for i in range(num_workers):
        for j in range(num_tasks):
            x[i,j]=solver.IntVar(0,1,"")
    #chaque machiniste a au plus 1 tâche
    for i in range (num_workers):
        solver.Add(solver.Sum([x[i,j] for j in range(num_tasks)]) <=1)
    #chaque tache est assignée a exactement un machiniste
    for j in range(num_tasks):
        solver.Add(solver.Sum([x[i,j] for i in range(num_workers)]) <= 1)
    #chaque machiniste ne peut que faire les tâches pour lesquelles il est accrédité
    for i in range(num_workers):
        for j in range(num_tasks):
            if D[i][j]==0:
                solver.Add(x[i,j]==0)
    objective_terms=[]
    for i in range(num_workers):
        for j in range(num_tasks):
            objective_terms.append(costs[i][j]*x[i,j])
    solver.Maximize(solver.Sum(objective_terms))

    status=solver.Solve()
    if status == pywraplp.Solver.OPTIMAL or status == pywraplp.Solver.FEASIBLE:
        print(f"Total cost = {solver.Objective().Value()}\n")
        for i in range(num_workers):
            for j in range(num_tasks):
                # Test if x[i,j] is 1 (with tolerance for floating point arithmetic).
                if x[i, j].solution_value() > 0.5:
                    y[i,j]=1
                    print(f'machiniste{i} fait le service{j}')
                    #print(f"Worker {i} assigned to task {j}." + f" Cost: {costs[i][j]}")
    #else:
        #print("No solution found.")
    return y
#x=opti()
def matrice_vers_dataframe(x, num_workers, num_tasks, identifiants, services):  
    #affectations = [[int(x[i, j].solution_value() > 0.5) for j in range(num_tasks)]for i in range(num_workers)]  
    df = pd.DataFrame(x, index=identifiants, columns=services)
    print (df) 
    return df


identifiants = pd.read_excel("Partie_1_LLM/data/Export_Planning_du_12_01_2026_au_16_01_2026.xlsx")['Identifiant'].tolist() 
services = pd.read_excel('Partie_1_LLM/data/Services_Agents_non_affectés_le_12_01_2026.xlsx')['Service'].tolist()
num_workers=len(D)
num_tasks=len(D[0])
df = matrice_vers_dataframe(opti(D,W), num_workers, num_tasks, identifiants, services)

df.to_excel("resultats12.xlsx")

#on va retrouver qui fait des services de nuit et les aprèms
# on cherche les services de nuit et de l'aprèm

def tri_horaire(chemin_fichier_serv):
    df_serv = pd.read_excel(chemin_fichier_serv)[['Service', 'Début', 'Fin', 'Type']]
    
    # Conversion propre pour les comparaisons
    df_serv['t_debut'] = df_serv['Début'].apply(to_time)
    df_serv['t_fin'] = df_serv['Fin'].apply(to_time)
    
    t05 = to_time("05:00")
    t11 = to_time("11:00")
    t14 = to_time("14:00")
    t22 = to_time("22:00")

    matin = df_serv[(df_serv['t_fin'] <= t14) & (df_serv['t_debut'] < t14)]
    aprem = df_serv[(df_serv['t_fin'] <= t22) & (df_serv['t_debut'] >= t14)]
    nuit = df_serv[(df_serv['t_fin'] > t22) | (df_serv['t_fin'] < t05)]
    coupure = df_serv[df_serv['Type'] == 'COUP']
    mixte = df_serv[(df_serv['t_fin'] > t14) & (df_serv['t_debut'] < t11)]
    
    return (matin, aprem, nuit, coupure, mixte)

'''
def correction_en_fonction_du_jour_d_avant(df_travail_veille):
    # 1. Initialisation des données
    D1 = initialize_data("Partie_1_LLM/data/Export_Planning_du_12_01_2026_au_16_01_2026.xlsx",
                         'Partie_1_LLM/data/Services_Agents_non_affectés_le_13_01_2026.xlsx','13/01/2026')
    serv = pd.read_excel('Partie_1_LLM/data/Services_Agents_non_affectés_le_13_01_2026.xlsx')
    
    D = matrice_vers_dataframe(D1, num_workers, num_tasks, identifiants, serv)
    
    tri_ajd = tri_horaire('Partie_1_LLM/data/Services Agents non affectés le 13_01_2026.xlsx')
    tri_hier = tri_horaire('Partie_1_LLM/data/Services Agents non affectés le 12_01_2026.xlsx')
    
    # 2. Parcours des agents de la veille
    for i in df_travail_veille['identifiants']:
        # On suppose que l'identifiant est l'index de df_travail_veille
        # et qu'il y a une colonne 'service'
        if i in df_travail_veille.index:
            test = df_travail_veille.loc[i, 'service']
        else:
            continue # Si l'identifiant n'est pas trouvé, on passe au suivant
            
        # Conversion des services en listes/sets pour accélérer la recherche avec 'in'
        services_hier_1 = tri_hier[1]['Service'].values
        services_hier_2 = tri_hier[2]['Service'].values
        
        # Conditions et modifications directes dans D
        # Remplacer le bloc de filtrage par :
        if test in services_hier_1:
            services_a_bloquer = pd.concat([tri_ajd[0]['Service'], tri_ajd[3]['Service']]).unique()
            # On filtre les colonnes existantes dans D qui sont dans 'services_a_bloquer'
            cols = [c for c in services_a_bloquer if c in D.columns]
            D.loc[i, cols] = 0
            
        elif test in services_hier_2:
            services_a_bloquer = pd.concat([tri_ajd[0]['Service'], tri_ajd[3]['Service'], tri_ajd[4]['Service']])
            mask = (D['Service'].isin(services_a_bloquer)) & (D['identifiant'] == i)
            D.loc[mask, :] = 0
            
    return D.to_numpy()
'''
def correction_en_fonction_du_jour_d_avant(df_travail_veille, num_workers, num_tasks, identifiants,chemin_ajd,chemin_hier,ajd):
    # 1. Initialisation des données pour le jour J (13/01/2026) -> AJOUT DES UNDERSCORES ICI
    D1 = initialize_data("Partie_1_LLM/data/Export_Planning_du_12_01_2026_au_16_01_2026.xlsx",
                         chemin_ajd, ajd)
    
    # Extraction de la liste propre des services pour aujourd'hui
    serv_df = pd.read_excel(chemin_ajd)
    serv_list = serv_df['Service'].tolist()
    
    # Reconstruction de la matrice D de base pour aujourd'hui
    D = matrice_vers_dataframe(D1, num_workers, num_tasks, identifiants, serv_list)
    
    # Récupération des tris horaires (Aujourd'hui vs Veille)
    tri_ajd = tri_horaire(chemin_ajd)
    tri_hier = tri_horaire(chemin_hier)
    
    # 2. Parcours des agents à partir du DataFrame de la veille
    for i in df_travail_veille.index:
        if i not in D.index:
            continue
            
        # Récupération sécurisée du service effectué hier par l'agent i
        test = df_travail_veille.loc[i, 'service'] if 'service' in df_travail_veille.columns else None
        if pd.isna(test) or test is None:
            continue
            
        services_hier_1 = tri_hier[1]['Service'].values  # Services Après-midi d'hier
        services_hier_2 = tri_hier[2]['Service'].values  # Services Nuit d'hier
        
        # Si l'agent a travaillé d'après-midi la veille
        if test in services_hier_1:
            services_a_bloquer = pd.concat([tri_ajd[0]['Service'], tri_ajd[3]['Service']]).unique()
            cols = [c for c in services_a_bloquer if c in D.columns]
            D.loc[i, cols] = 0
            
        # Si l'agent a travaillé de nuit la veille
        elif test in services_hier_2:
            services_a_bloquer = pd.concat([tri_ajd[0]['Service'], tri_ajd[3]['Service'], tri_ajd[4]['Service']]).unique()
            cols = [c for c in services_a_bloquer if c in D.columns]
            D.loc[i, cols] = 0
            
    return D.to_numpy()
 #on fait le 13             
D13=correction_en_fonction_du_jour_d_avant(df,len(D),len((pd.read_excel('Partie_1_LLM/data/Services_Agents_non_affectés_le_13_01_2026.xlsx')) [['Service', 'Début', 'Fin']]),identifiants,'Partie_1_LLM/data/Services_Agents_non_affectés_le_13_01_2026.xlsx','Partie_1_LLM/data/Services_Agents_non_affectés_le_12_01_2026.xlsx','13/01/2026')
W13 = W_initialize("Partie_2_Optimisation/preferences_agents.xlsx", 'Partie_1_LLM/data/Services_Agents_non_affectés_le_13_01_2026.xlsx', (len(D13), len(D13[0])), D13)
services13 = pd.read_excel('Partie_1_LLM/data/Services_Agents_non_affectés_le_13_01_2026.xlsx')['Service'].tolist()
print("jour13")
df13 = matrice_vers_dataframe(opti(D13,W13), num_workers, num_tasks, identifiants, services13)
df13.to_excel("resultats13.xlsx")

#on fait le 14

D14=correction_en_fonction_du_jour_d_avant(df,len(D),len((pd.read_excel('Partie_1_LLM/data/Services_Agents_non_affectés_le_14_01_2026.xlsx')) [['Service', 'Début', 'Fin']])),identifiants,'Partie_1_LLM/data/Services_Agents_non_affectés_le_14_01_2026.xlsx','Partie_1_LLM/data/Services_Agents_non_affectés_le_13_01_2026.xlsx','14/01/2026')
W14 = W_initialize("Partie_2_Optimisation/preferences_agents.xlsx", 'Partie_1_LLM/data/Services_Agents_non_affectés_le_13_01_2026.xlsx', (len(D14), len(D14[0])), D14)
services14 = pd.read_excel('Partie_1_LLM/data/Services_Agents_non_affectés_le_14_01_2026.xlsx')['Service'].tolist()
df14 = matrice_vers_dataframe(opti(D14,W14), num_workers, len(D15[0]), identifiants, services14)
df14.to_excel("resultats14.xlsx")

#on fait le 15

D15=correction_en_fonction_du_jour_d_avant(df,len(D),len((pd.read_excel('Partie_1_LLM/data/Services_Agents_non_affectés_le_15_01_2026.xlsx')) [['Service', 'Début', 'Fin']])),identifiants,'Partie_1_LLM/data/Services_Agents_non_affectés_le_15_01_2026.xlsx','Partie_1_LLM/data/Services_Agents_non_affectés_le_14_01_2026.xlsx','15/01/2026')
W15 = W_initialize("Partie_2_Optimisation/preferences_agents.xlsx", 'Partie_1_LLM/data/Services_Agents_non_affectés_le_15_01_2026.xlsx', (len(D15), len(D15[0])), D15)
services15 = pd.read_excel('Partie_1_LLM/data/Services_Agents_non_affectés_le_15_01_2026.xlsx')['Service'].tolist()
df15 = matrice_vers_dataframe(opti(D15,W15), num_workers,len(D15[0]) , identifiants, services15)
df15.to_excel("resultats14.xlsx")

#on fait le 16


D16=correction_en_fonction_du_jour_d_avant(df,len(D),len((pd.read_excel('Partie_1_LLM/data/Services_Agents_non_affectés_le_16_01_2026.xlsx')) [['Service', 'Début', 'Fin']])),identifiants,'Partie_1_LLM/data/Services_Agents_non_affectés_le_16_01_2026.xlsx','Partie_1_LLM/data/Services_Agents_non_affectés_le_15_01_2026.xlsx','16/01/2026')
W16 = W_initialize("Partie_2_Optimisation/preferences_agents.xlsx", 'Partie_1_LLM/data/Services_Agents_non_affectés_le_16_01_2026.xlsx', (len(D16), len(D16[0])), D16)
services16 = pd.read_excel('Partie_1_LLM/data/Services_Agents_non_affectés_le_16_01_2026.xlsx')['Service'].tolist()
df16 = matrice_vers_dataframe(opti(D16,W16), num_workers, len(D16[0]), identifiants, services16)
df16.to_excel("resultats16.xlsx")

        


    
res = opti(W,D)

#On crée une fonction qui va update le planning excel de la semaine
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def update_planning(res,day):
    df_serv = (pd.read_excel(f"Partie_1_LLM/data/Services_Agents_non_affectés_le_{day.replace("/","_")}.xlsx")) [['Service', 'Début', 'Fin']]
    wb = load_workbook("Partie_1_LLM/data/Export_Planning_du_12_01_2026_au_16_01_2026.xlsx")
    ws = wb.active
    numero_colonne = None

    for cell in ws[1]:  # ligne 1 = ligne des titres
        if cell.value == day:
            numero_colonne = cell.column
            break

    vert = PatternFill(
        start_color="92D050",
        end_color="92D050",
        fill_type="solid"
    )

    rouge = PatternFill(
        start_color="FF0000",
        end_color="FF0000",
        fill_type="solid"
    )
    jaune_pastel = PatternFill(
    start_color="FFF2CC",
    end_color="FFF2CC",
    fill_type="solid"
    )

    for i in range(len(D)):
        if np.any(res[i,:]):
            index_serv = np.where(res[i,:])
            service_name = df_serv['Service'].iloc[index_serv]
            cellule= ws.cell(row=i + 2, column=numero_colonne)
            cellule.value= (str(service_name))[7:7+8]
            cellule.fill = jaune_pastel

    wb.save(f"planning_du_{day.replace("/","_")}_updated.xlsx")


def create_dico_affectés(mat_res, day):
    ans = []
    df_serv = (pd.read_excel(f"Partie_1_LLM/data/Services_Agents_non_affectés_le_{day.replace("/","_")}.xlsx")) [['Service', 'Début', 'Fin']]
    df_mach =  pd.read_excel("Partie_1_LLM/data/Export_Planning_du_12_01_2026_au_16_01_2026.xlsx")['Identifiant']
    N = len(df_mach)
    for i in range(N):
        if np.any(mat_res[i, :]):
            j = np.where(mat_res[i,:] == 1)
            ans.append({"agent": str(df_mach['Identifiant'].iloc[i]), "service": str(df_serv['Service'].iloc[j])})
    return ans

def create_liste_non_affecté(mat_res, day):
    ans = []
    df_serv = (pd.read_excel(f"Partie_1_LLM/data/Services_Agents_non_affectés_le_{day.replace("/","_")}.xlsx")) [['Service', 'Début', 'Fin']]
    P = len(df_serv)
    for j in range(P):
        if not np.any(mat_res[:,j]):
            ans.append(df_serv['Service'].iloc[j])
    return ans







